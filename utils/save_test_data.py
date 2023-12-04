import openpyxl
import os
from openpyxl.styles import Side, Border

def save_data(data_dict):

    folder_path = "./result"
    data_file = 'test_data.xlsx'

    if not os.path.exists(folder_path):
      os.mkdir(folder_path)

    # 检查文件夹中是否有data_file文件
    if data_file in os.listdir(folder_path):
        # 如果有，则打开该文件
        workbook = openpyxl.load_workbook(os.path.join(folder_path, data_file))
        worksheet = workbook.active
    else:
        # 如果没有，则新建一个Excel文件
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

    print('save test result...')

    # 设置行高
    worksheet.row_dimensions[1].height = 20

    # 设置列宽
    worksheet.column_dimensions['A'].width = 16
    worksheet.column_dimensions['B'].width = 16
    worksheet.column_dimensions['C'].width = 16
    worksheet.column_dimensions['D'].width = 16
    worksheet.column_dimensions['E'].width = 16
    worksheet.column_dimensions['F'].width = 16
    worksheet.column_dimensions['G'].width = 16
    worksheet.column_dimensions['H'].width = 16
    worksheet.column_dimensions['I'].width = 16
    worksheet.column_dimensions['J'].width = 16
    worksheet.column_dimensions['K'].width = 16
    worksheet.column_dimensions['L'].width = 16
    worksheet.column_dimensions['M'].width = 16

    # 设置第1行第2列到第5列合并单元格，并写入Accuray
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    worksheet.cell(row=1, column=2, value="Accuray").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 设置第1行第6列到第9列合并单元格，并写入PSNR
    worksheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)
    worksheet.cell(row=1, column=6, value="PSNR").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 设置第1行第10列到第13列合并单元格，并写入SSIM
    worksheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)
    worksheet.cell(row=1, column=10, value="SSIM").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 写入easy_acc，medium_acc，hard_acc，avg_acc
    worksheet.cell(row=2, column=2, value="easy_acc").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=3, value="medium_acc").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=4, value="hard_acc").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=5, value="avg_acc").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 写入easy_psnr，medium_psnr，hard_psnr，avg_psnr
    worksheet.cell(row=2, column=6, value="easy_psnr").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=7, value="medium_psnr").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=8, value="hard_psnr").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=9, value="avg_psnr").alignment = worksheet.cell(row=2, column=10, value="easy_ssim").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=11, value="medium_ssim").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=12, value="hard_ssim").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=13, value="avg_ssim").alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    if data_dict["rec"]=='ASTER':
        
        # 写入rec，easy_acc，medium_acc，hard_acc，avg_acc的值
        worksheet.cell(row=3, column=1, value=data_dict["rec"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=2, value=data_dict["easy_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=3, value=data_dict["medium_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=4, value=data_dict["hard_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=5, value=data_dict["avg_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 写入easy_psnr，medium_psnr，hard_psnr，avg_psnr的值
        worksheet.cell(row=3, column=6, value=data_dict["easy_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=7, value=data_dict["medium_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=8, value=data_dict["hard_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=9, value=data_dict["avg_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 写入easy_ssim，medium_ssim，hard_ssim，avg_ssim的值
        worksheet.cell(row=3, column=10, value=data_dict["easy_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=11, value=data_dict["medium_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=12, value=data_dict["hard_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=3, column=13, value=data_dict["avg_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    if data_dict["rec"]=='MORAN':
        
        # 写入rec，easy_acc，medium_acc，hard_acc，avg_acc的值
        worksheet.cell(row=4, column=1, value=data_dict["rec"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=2, value=data_dict["easy_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=3, value=data_dict["medium_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=4, value=data_dict["hard_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=5, value=data_dict["avg_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 写入easy_psnr，medium_psnr，hard_psnr，avg_psnr的值
        worksheet.cell(row=4, column=6, value=data_dict["easy_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=7, value=data_dict["medium_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=8, value=data_dict["hard_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=9, value=data_dict["avg_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 写入easy_ssim，medium_ssim，hard_ssim，avg_ssim的值
        worksheet.cell(row=4, column=10, value=data_dict["easy_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=11, value=data_dict["medium_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=12, value=data_dict["hard_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=4, column=13, value=data_dict["avg_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')    

    if data_dict["rec"]=='CRNN':
        
        # 写入rec，easy_acc，medium_acc，hard_acc，avg_acc的值
        worksheet.cell(row=5, column=1, value=data_dict["rec"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=2, value=data_dict["easy_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=3, value=data_dict["medium_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=4, value=data_dict["hard_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=5, value=data_dict["avg_acc"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 写入easy_psnr，medium_psnr，hard_psnr，avg_psnr的值
        worksheet.cell(row=5, column=6, value=data_dict["easy_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=7, value=data_dict["medium_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=8, value=data_dict["hard_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=9, value=data_dict["avg_psnr"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # 写入easy_ssim，medium_ssim，hard_ssim，avg_ssim的值
        worksheet.cell(row=5, column=10, value=data_dict["easy_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=11, value=data_dict["medium_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=12, value=data_dict["hard_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=13, value=data_dict["avg_ssim"]).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 设置表格字体
    font_bold = openpyxl.styles.Font(name="Arial", size=16, bold=True)
    font_regular = openpyxl.styles.Font(name="Arial", size=14, bold=False)

    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=13):
        for cell in row:
            cell.font = font_bold

    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=13):
        for cell in row:
            cell.font = font_regular

    # 加粗第一列
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = font_bold
            
    # 设置右边框线的样式
    border_right = Border(right=Side(style='thick'))

    # 遍历表格每行，对需要设置右边框线的单元格设置边框
    for row in worksheet.iter_rows():
        row[0].border = border_right  # 第1列
        row[4].border = border_right  # 第5列
        row[8].border = border_right  # 第9列
        row[12].border = border_right  # 第13列
        
    # 保存Excel文件
    workbook.save(os.path.join(folder_path, data_file))
    print('save test result success')

